#接口自动化步骤
#1.准备好接口测试用例
#2、获取接口测试用例
#3、提交接口请求
#4、获取接口返回的响应结果，并添加断言
#5、把结果写进表格里面

#导入相关库
import openpyxl
import requests


#读取接口测试用例，封装函数
def read_fuc(filename,sheetname):
    web=openpyxl.load_workbook(filename)  #打开工作簿
    sheet_login=web[sheetname]  #找到表单
    #获取到当前表单里面的最大行数  sheetname.max_row
    rows1=sheet_login.max_row
    login_list=[] #定义一个空列表，用来存储字典
    for i in range(2,rows1+1): #从第二行开始读取，range取头不取尾，要记得加1
        #因为获取出来的数据太散了，所以可以用字典对它们进行存取起来
        #可以把字典放在一个列表里面，方便后期的获取
        dict_data=dict(
        login_id=sheet_login.cell(row=i,column=1).value,  #找到相应表格内容并打印
        login_url = sheet_login.cell(row=i, column=6).value,  # 找到相应表格内容并打印
        login_data = sheet_login.cell(row=i, column=7).value,  # 找到相应表格内容并打印
        login_result = sheet_login.cell(row=i, column=8).value  # 找到相应表格内容并打印
        )
        login_list.append(dict_data)
    return login_list  #返回测试用例列表
result_web=read_fuc(filename='接口测试用例.xlsx', sheetname='login')
# print(result_web)

#发送接口请求
def request_fuc(l_url,l_ldata):
    header={'X-Lemonban-Media-Type':'lemonban.v2',   #请求头部信息
           'Content-Type':'application/json'}
    result=requests.post(url=l_url,json=l_ldata,headers=header)  #返回响应结果
    result1=result.json()
    return result1
#把结果写入表格
#打开工作簿
def write_result(filename,sheetname,rows,column,pass_res):
    wb=openpyxl.load_workbook(filename)
    #找到表单
    sheet=wb[sheetname]
    #找到单元格
    sheet.cell(row=rows,column=column).value=pass_res
    #保存
    wb.save('接口测试用例模板参考.xlsx')


#获取测试用例返回的结果-预期结果、执行结果-加断言
login_tall=read_fuc('接口测试用例模板参考.xlsx','login')  #调用读取的函数
for item in login_tall:   #遍历列表中的数据
    login_id=item.get('login_id') #取到id
    login_url = item.get('login_url') #取到url
    login_data = item['login_data']
    login_data=eval(login_data)  # 运行被字符串包裹的python表达式（外面有个引号）
    expect=item.get('login_result')     #取到预期结果
    expect=eval(expect)       #把字符串转换成字典
    expect_msg=expect.get('msg')    #取到预期结果里面的msg
    real_result=request_fuc(l_url=login_url,l_ldata=login_data)    #调用请求接口函数，并进行传参,并返回一个结果
    real_msg=real_result.get('msg')   #取到执行结果里面的msg
    print("预期结果为：{}".format(expect_msg))
    print("预期结果为：{}".format(real_msg))
    if real_msg==expect_msg:
        print("此用例通过测试")
        pass_result='passed'#变量用来标识最终结果
    else:
        print("此用例测试不通过")
        pass_result='failsed'
    write_result('接口测试用例模板参考.xlsx','login',login_id+1,10,pass_result)  #调用函数，并把结果写入表格



