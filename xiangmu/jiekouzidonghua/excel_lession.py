# 注意：excel表格放到和python文件同一级
# excel表格三大对象：
# 1、工作簿对象
# 2、表格对象：sheet
# 3、获取单元格对象，要得到单元格里面的值的话要用value


#写入数据：
# 1、对value进行赋值
# 2、保存--只有保存后才算修改成功
#导入openpyxl库
import openpyxl

def case_fuc(filename,sheetname):
    #获取工作簿，excel表格名字
    web=openpyxl.load_workbook(filename)
    #获取表单
    sheet_login=web[sheetname]
    #for循环找出表单中所有的行与列
    #web.max_row获取最大的行数
    number=sheet_login.max_row
    case_list=[]  #空列表存放测试用例
    for i in range(2,number+1):  #从第二行开始。。。取头不取尾
        #一、获取单元格,通过表单获取行 列号
        # 二、因为取出来的数据时一个一个的，所以可以把它们打个包，用字典装起来，上面的格式好像对象的属性
       dict1= dict(
        case_id=sheet_login.cell(row=i,column=1).value,      #获取case_id的值
        title=sheet_login.cell(row=i,column=4).value,       #获取title的值
        test_data=sheet_login.cell(row=i,column=6).value)    #获取url的值
       case_list.append(dict1)   #不能进行重新赋值操作，否则结果会返回None
    return case_list

result=case_fuc(filename='testcase_web.xlsx',sheetname='login')  #调用函数，传入参数
print(result)

#写入结果
wb=openpyxl.load_workbook('testcase_web.xlsx')  #加载工作簿
biaodan=wb['login']#打开表单
biaodan.cell(row=1,column=8).value='passed'#找到单元格并进行结果的写入
wb.save()#保存表格





# #对这个单元格内容进行重新赋值，只能直接对value进行赋值，否则会修改不成功--写入新数据到工作簿里面
# login_cell.value='用例编号1'
# print(result)
# #执行保存操作，记得要先关闭工作簿，否则会报错。可以直接保存，也可以进行另存为
# web.save('testcase_web.xlsx')