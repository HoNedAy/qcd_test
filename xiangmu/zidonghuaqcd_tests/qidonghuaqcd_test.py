#测试相关的内容

import requests #导入包
import openpyxl
# #打开文档
def open_execl(filename,sheetname):
    wb=openpyxl.load_workbook(filename) #找到文档
    sheeet=wb[sheetname]#找到表单
    num=sheeet.max_row
    list=[]
    for i in range(2,num+1):
        dict_data = dict(
        id=sheeet.cell(row=i, column=1).value,
        url=sheeet.cell(row=i, column=6).value,
        data = sheeet.cell(row=i, column=7).value,
        expect_result=sheeet.cell(row=i, column=8).value
        )
        list.append(dict_data)
    return list  #得到excel里面的数据
excel_result=open_execl('接口测试用例模板参考.xlsx','login')

#发送请求
def request_post(url,data):
    header = {'X-Lemonban-Media-Type': 'lemonban.v2',  # 请求头部信息
              'Content-Type': 'application/json'}
    result=requests.post(url=url,json=data,headers=header)  #返回响应结果
    return  result

def write_result(filanema,sheetname,row,column,pass_result):
    we=openpyxl.load_workbook(filanema)
    sheet=we[sheetname]
    sheet.cell(row=row,column=column).value=pass_result #找到单元格的列
    we.save('接口测试用例模板参考.xlsx')#保存



#获取测试用例返回的结果-预期结果、执行结果-加断言
open_execl_1=open_execl('接口测试用例模板参考.xlsx','login')
for item in open_execl_1:
    id=item.get('id')
    url=item.get('url')
    data=item['data']
    data=eval(data)
    expect=item.get('expect_result')
    expect=eval(expect)
    expect_result=expect.get('msg')
    real_selut=request_post(url=url,data=data)
    dd=real_selut.json()
    real_selut_data=dd.get('msg')
    print(expect_result+'-------'+real_selut_data)
    if expect_result==real_selut_data:
        print('测试通过，结果是'+real_selut_data)
        pass_result='true'
    else:
        print('测试不通过，结果是'+real_selut_data)
        pass_result='false'
    #把结果写入文档
    write_result('接口测试用例模板参考.xlsx','login',id+1,9,pass_result)




